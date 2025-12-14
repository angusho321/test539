#!/usr/bin/env python3
"""
週一冠軍策略分析腳本
將分析結果寫入 Excel 的 Monday_Strategy 分頁
"""

import pandas as pd
import numpy as np
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
import sys

def load_lottery_data(file_path):
    """讀取彩票歷史資料"""
    if not Path(file_path).exists():
        print(f"❌ 檔案不存在: {file_path}")
        return None
    
    try:
        df = pd.read_excel(file_path, engine='openpyxl', sheet_name='Sheet1')
    except:
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
        except Exception as e:
            print(f"❌ 讀取 Excel 失敗: {e}")
            return None
    
    # 處理日期欄位
    if '日期' in df.columns:
        try:
            df['日期'] = pd.to_datetime(df['日期'], format='mixed', errors='coerce')
            if df['日期'].isna().any():
                df['日期'] = pd.to_datetime(df['日期'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
            if df['日期'].isna().any():
                df['日期'] = pd.to_datetime(df['日期'], format='%Y-%m-%d', errors='coerce')
        except:
            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
        
        df = df.dropna(subset=['日期'])
        df = df.sort_values('日期').reset_index(drop=True)
    
    return df

def get_monday_records(df):
    """取得所有週一的開獎記錄"""
    if df is None or df.empty:
        return pd.DataFrame()
    
    # 確保日期是 datetime
    df['weekday'] = df['日期'].dt.weekday  # 0=週一, 6=週日
    monday_records = df[df['weekday'] == 0].copy()
    return monday_records.sort_values('日期').reset_index(drop=True)

def calculate_number_with_offset(base_number, offset):
    """
    計算新號碼：基準號碼 + 加數
    特殊處理：若相加結果 > 39，則結果 = 結果 - 39
    """
    result = base_number + offset
    if result > 39:
        result = result - 39
    return result

def calculate_strategy_numbers(monday_record, ball_a_index, ball_b_index, offset_a, offset_b):
    """
    根據週一開獎記錄計算策略號碼
    ball_a_index: 第一顆球的索引（1-5，對應號碼1-號碼5）
    ball_b_index: 第二顆球的索引（1-5，對應號碼1-號碼5）
    offset_a: 第一顆球的偏移量
    offset_b: 第二顆球的偏移量
    """
    if offset_a is None or offset_b is None:
        raise ValueError("offset_a 和 offset_b 必須提供，不能為 None")
    
    # 取得對應的球號
    num_a = int(monday_record[f'號碼{ball_a_index}'])
    num_b = int(monday_record[f'號碼{ball_b_index}'])
    
    A = calculate_number_with_offset(num_a, offset_a)
    B = calculate_number_with_offset(num_b, offset_b)
    
    return A, B

def get_target_weekdays(lottery_type):
    """取得目標追號期
    - 539: 週二至週六 (1, 2, 3, 4, 5)
    - Fantasy5: 週二至週日 (1, 2, 3, 4, 5, 6) - 因為天天樂週一到週日都有開獎
    """
    if lottery_type == 'fantasy5':
        return [1, 2, 3, 4, 5, 6]  # 週二至週日
    else:
        return [1, 2, 3, 4, 5]  # 週二至週六

def preprocess_weekly_data(df, monday_records, lottery_type, weeks=52):
    """
    預先處理資料，建立每週的資料結構
    返回: List[Dict]，每個元素包含該週一的5顆球號碼和該週目標日期所有開出的號碼 Set
    - Fantasy5: 週二至週日
    - 539: 週二至週六
    """
    # 只取最近 N 週的週一記錄
    recent_mondays = monday_records.tail(weeks).copy()
    
    if recent_mondays.empty:
        return []
    
    target_weekdays = get_target_weekdays(lottery_type)  # 根據彩種決定範圍
    weekly_data = []
    
    for idx, monday_row in recent_mondays.iterrows():
        monday_date = monday_row['日期']
        
        # 取得該週一的5顆球號碼
        monday_nums = [
            int(monday_row['號碼1']),
            int(monday_row['號碼2']),
            int(monday_row['號碼3']),
            int(monday_row['號碼4']),
            int(monday_row['號碼5'])
        ]
        
        # 找出這個週一之後的目標日期開獎記錄（只查詢一次）
        # Fantasy5: 週二至週日
        # 539: 週二至週六
        # 使用日期（不含時間）來比較，避免時間戳記造成的問題
        if hasattr(monday_date, 'date'):
            monday_date_only = monday_date.date()
        elif isinstance(monday_date, pd.Timestamp):
            monday_date_only = monday_date.date()
        else:
            monday_date_only = monday_date
        
        week_start = pd.Timestamp(monday_date_only) + timedelta(days=1)  # 週二 00:00:00
        # 根據彩種決定結束日期
        if lottery_type == 'fantasy5':
            week_end = pd.Timestamp(monday_date_only) + timedelta(days=6)  # 週日 00:00:00
        else:
            week_end = pd.Timestamp(monday_date_only) + timedelta(days=5)  # 週六 00:00:00 (539)
        
        # 過濾：日期在週二至目標結束日期之間，且 weekday 符合目標範圍
        week_records = df[
            (df['日期'] >= week_start) & 
            (df['日期'] <= week_end) &
            (df['日期'].dt.weekday.isin(target_weekdays))
        ].copy()
        
        # 建立該週所有開出號碼的 Set（用於快速查找）
        winning_set = set()
        # 儲存每一天的開獎記錄（按日期排序，用於統計每一天的中獎情況）
        daily_records = []  # List of (weekday, drawn_numbers)
        
        if not week_records.empty:
            # 按日期排序
            week_records_sorted = week_records.sort_values('日期')
            for _, record in week_records_sorted.iterrows():
                drawn_numbers = [
                    int(record['號碼1']),
                    int(record['號碼2']),
                    int(record['號碼3']),
                    int(record['號碼4']),
                    int(record['號碼5'])
                ]
                winning_set.update(drawn_numbers)
                # 儲存每一天的記錄（weekday: 1=週二, 2=週三, ..., 5=週六, 6=週日）
                weekday = record['日期'].weekday()
                daily_records.append((weekday, drawn_numbers))
        
        weekly_data.append({
            'monday_date': monday_date,  # 保存週一日期，用於顯示
            'monday_nums': monday_nums,
            'winning_set': winning_set,
            'daily_records': daily_records,  # 每一天的開獎記錄
            'has_data': len(winning_set) > 0  # 標記是否有開獎資料
        })
    
    return weekly_data

def backtest_strategy_optimized(weekly_data, ball_a_index, ball_b_index, offset_a, offset_b):
    """
    優化版回測策略：使用預處理的資料進行純記憶體比對
    weekly_data: 預處理的每週資料（來自 preprocess_weekly_data）
    ball_a_index: 第一顆球的索引（1-5）
    ball_b_index: 第二顆球的索引（1-5）
    offset_a: 第一顆球的偏移量
    offset_b: 第二顆球的偏移量
    
    返回: (win_rate, wins, total, missed_weeks, day_stats)
    missed_weeks: 未中獎的週一日期列表
    day_stats: 字典，記錄每一天的中獎次數 {1: count, 2: count, ...} (1=週二, 2=週三, ...)
    """
    if not weekly_data:
        # 如果沒有資料，預設返回不包含週日的格式（539）
        return 0.0, 0, 0, [], {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    
    wins = 0
    total = 0
    missed_weeks = []  # 記錄未中獎的週
    # 根據 weekly_data 的第一筆記錄判斷是否包含週日
    has_sunday = len(weekly_data) > 0 and any(6 in [r[0] for r in week_info.get('daily_records', [])] for week_info in weekly_data)
    if has_sunday:
        day_stats = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0}  # 1=週二, 2=週三, 3=週四, 4=週五, 5=週六, 6=週日
    else:
        day_stats = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}  # 1=週二, 2=週三, 3=週四, 4=週五, 5=週六
    
    for week_info in weekly_data:
        # 跳過沒有開獎資料的週
        if not week_info['has_data']:
            continue
        
        monday_nums = week_info['monday_nums']
        monday_date = week_info['monday_date']
        
        # 計算策略號碼（純記憶體運算）
        num_a = monday_nums[ball_a_index - 1]  # 轉換為 0-based 索引
        num_b = monday_nums[ball_b_index - 1]
        
        A = calculate_number_with_offset(num_a, offset_a)
        B = calculate_number_with_offset(num_b, offset_b)
        
        total += 1
        
        # 使用預處理的 daily_records 來統計每一天的中獎情況（避免查詢 DataFrame）
        daily_records = week_info.get('daily_records', [])
        winning_set = week_info['winning_set']
        
        if A in winning_set or B in winning_set:
            wins += 1
            # 找出第一次中獎的日期（用於統計每一天的中獎情況）
            found_win = False
            for weekday, drawn_numbers in daily_records:
                if A in drawn_numbers or B in drawn_numbers:
                    day_stats[weekday] += 1
                    found_win = True
                    break  # 只記錄第一次中獎
        else:
            # 記錄未中獎的週
            missed_weeks.append(monday_date)
    
    win_rate = (wins / total * 100) if total > 0 else 0.0
    return win_rate, wins, total, missed_weeks, day_stats

def find_best_strategies(df, monday_records, lottery_type, weeks=52, min_win_rate=90.0):
    """
    動態分析過去一年的歷史數據，找出勝率超過指定閾值的最佳策略組合
    測試所有可能的球號組合（第1-5支）和所有偏移量組合（0-38）
    返回前兩名最佳策略
    
    優化：使用預先計算的資料結構，避免在迴圈內查詢 DataFrame
    """
    if monday_records.empty or len(monday_records) < 2:
        return []
    
    print(f"🔍 開始動態分析所有可能的策略組合...")
    print(f"   球號組合: 第1-5支 × 第1-5支 = 25 種")
    print(f"   偏移量組合: 0-38 × 0-38 = 1521 種")
    print(f"   總組合數: 25 × 1521 = 38025 種")
    
    # Step 1: 預先處理資料（只執行一次）
    print(f"   📊 預先處理資料中...")
    weekly_data = preprocess_weekly_data(df, monday_records, lottery_type, weeks)
    
    if not weekly_data:
        return []
    
    print(f"   ✅ 資料預處理完成，共 {len(weekly_data)} 週資料")
    
    # Step 2: 嘗試所有可能的球號組合（1-5）和偏移量組合（0-38）
    all_strategies = []
    total_combinations = 5 * 5 * 39 * 39  # 5×5×39×39 = 38025 種組合
    processed = 0
    
    print(f"   🚀 開始回測（純記憶體比對模式）...")
    
    for ball_a_index in range(1, 6):  # 第1支到第5支
        for ball_b_index in range(1, 6):  # 第1支到第5支
            for offset_a in range(0, 39):
                for offset_b in range(0, 39):
                    processed += 1
                    if processed % 5000 == 0:
                        progress = (processed / total_combinations) * 100
                        print(f"   進度: {progress:.1f}% ({processed}/{total_combinations})", end='\r', flush=True)
                    
                    # 回測這個策略組合（使用優化版函數，純記憶體比對）
                    win_rate, wins, total, missed_weeks, day_stats = backtest_strategy_optimized(
                        weekly_data,
                        ball_a_index, ball_b_index,
                        offset_a, offset_b
                    )
                    
                    # 只保留勝率超過閾值的策略
                    if win_rate >= min_win_rate and total > 0:
                        all_strategies.append({
                            'ball_a_index': ball_a_index,
                            'ball_b_index': ball_b_index,
                            'offset_a': offset_a,
                            'offset_b': offset_b,
                            'win_rate': win_rate,
                            'wins': wins,
                            'total': total,
                            'missed_weeks': missed_weeks,  # 記錄未中獎的週
                            'day_stats': day_stats  # 記錄每一天的中獎次數
                        })
    
    print(f"\n   完成！找到 {len(all_strategies)} 組勝率 >= {min_win_rate}% 的策略")
    
    # 排序：先按勝率降序，再按中獎次數降序
    all_strategies.sort(key=lambda x: (-x['win_rate'], -x['wins']))
    
    # 去除重複的策略組合（不考慮順序）
    # 例如：(第一顆球+6, 第二顆球+12) 和 (第二顆球+12, 第一顆球+6) 視為相同
    seen_strategies = set()
    unique_strategies = []
    
    for strategy in all_strategies:
        # 建立策略的唯一標識（標準化：較小的球號在前，如果球號相同則較小的偏移量在前）
        ball_a = strategy['ball_a_index']
        ball_b = strategy['ball_b_index']
        offset_a = strategy['offset_a']
        offset_b = strategy['offset_b']
        
        # 標準化：確保 (ball_a, offset_a) <= (ball_b, offset_b)
        if (ball_a, offset_a) > (ball_b, offset_b):
            # 交換順序
            strategy_key = ((ball_b, offset_b), (ball_a, offset_a))
        else:
            strategy_key = ((ball_a, offset_a), (ball_b, offset_b))
        
        # 如果這個策略組合還沒見過，加入結果
        if strategy_key not in seen_strategies:
            seen_strategies.add(strategy_key)
            unique_strategies.append(strategy)
    
    print(f"   去除重複後，剩餘 {len(unique_strategies)} 組唯一策略")
    
    # 返回前五名（已去重）
    return unique_strategies[:5]

def check_current_week_status(df, latest_monday, ball_a_index, ball_b_index, offset_a, offset_b, lottery_type):
    """檢查本週狀態（使用指定的球號和 offset）"""
    if latest_monday is None:
        return "無資料", None, None
    
    monday_date = latest_monday['日期']
    A, B = calculate_strategy_numbers(latest_monday, ball_a_index, ball_b_index, offset_a, offset_b)
    
    # 找出本週的目標日期開獎記錄
    # Fantasy5: 週二至週日
    # 539: 週二至週六
    # 使用日期（不含時間）來比較，避免時間戳記造成的問題
    if hasattr(monday_date, 'date'):
        monday_date_only = monday_date.date()
    elif isinstance(monday_date, pd.Timestamp):
        monday_date_only = monday_date.date()
    else:
        monday_date_only = monday_date
    
    week_start = pd.Timestamp(monday_date_only) + timedelta(days=1)  # 週二 00:00:00
    # 根據彩種決定結束日期
    if lottery_type == 'fantasy5':
        week_end = pd.Timestamp(monday_date_only) + timedelta(days=6)  # 週日 00:00:00
    else:
        week_end = pd.Timestamp(monday_date_only) + timedelta(days=5)  # 週六 00:00:00 (539)
    target_weekdays = get_target_weekdays(lottery_type)  # 根據彩種決定範圍
    
    week_records = df[
        (df['日期'] >= week_start) & 
        (df['日期'] <= week_end) &
        (df['日期'].dt.weekday.isin(target_weekdays))
    ].copy()
    
    if week_records.empty:
        return "等待開獎", None, None
    
    # 檢查是否已開獎
    today = datetime.now().date()
    latest_record_date = week_records['日期'].max().date()
    
    # 檢查是否中獎
    for _, record in week_records.iterrows():
        drawn_numbers = [
            int(record['號碼1']),
            int(record['號碼2']),
            int(record['號碼3']),
            int(record['號碼4']),
            int(record['號碼5'])
        ]
        
        if A in drawn_numbers or B in drawn_numbers:
            win_date = record['日期'].date()
            return "已中獎", win_date, record
    
    # 如果本週已過完但沒中獎
    # Fantasy5: 檢查到週日（+6天）
    # 539: 檢查到週六（+5天）
    if lottery_type == 'fantasy5':
        check_days = 6  # 週日
    else:
        check_days = 5  # 週六
    
    if latest_record_date >= (monday_date + timedelta(days=check_days)).date():
        return "未中獎", None, None
    
    return "等待開獎", None, None

def add_strategy_sheet(file_path, lottery_type):
    """將策略分析結果寫入 Excel 的新分頁"""
    print(f"📊 開始分析 {lottery_type} 的週一冠軍策略...")
    
    # 讀取資料
    df = load_lottery_data(file_path)
    if df is None:
        return False
    
    print(f"✅ 成功讀取 {len(df)} 筆歷史記錄")
    
    # 取得週一記錄
    monday_records = get_monday_records(df)
    if monday_records.empty:
        print("❌ 沒有找到週一的開獎記錄")
        return False
    
    print(f"📅 找到 {len(monday_records)} 筆週一開獎記錄")
    
    # 取得最新的週一記錄
    latest_monday = monday_records.iloc[-1]
    latest_monday_date = latest_monday['日期']
    
    # 動態分析找出最佳策略（勝率 > 90%）
    best_strategies = find_best_strategies(df, monday_records, lottery_type, weeks=52, min_win_rate=90.0)
    
    # 準備最佳策略字串
    first_strategy_str = "無符合策略"
    second_strategy_str = "無符合策略"
    
    # 球號中文對應
    ball_names = {1: '第一顆球', 2: '第二顆球', 3: '第三顆球', 4: '第四顆球', 5: '第五顆球'}
    
    def format_missed_weeks(missed_weeks):
        """格式化未中獎的週列表"""
        if not missed_weeks:
            return "無"
        # 格式化日期為 YYYY-MM-DD
        dates_str = ", ".join([date.strftime('%Y-%m-%d') for date in missed_weeks])
        return dates_str
    
    def format_day_stats(day_stats, total_wins):
        """格式化每一天的中獎統計"""
        if total_wins == 0:
            return "無中獎記錄"
        
        day_names = {1: '週二', 2: '週三', 3: '週四', 4: '週五', 5: '週六', 6: '週日'}
        day_labels = {1: '第一天', 2: '第二天', 3: '第三天', 4: '第四天', 5: '第五天', 6: '第六天'}
        
        # 根據 day_stats 中是否有週日來決定範圍
        has_sunday = 6 in day_stats and day_stats[6] > 0
        days_to_check = [1, 2, 3, 4, 5, 6] if has_sunday else [1, 2, 3, 4, 5]
        
        lines = []
        for day in days_to_check:
            count = day_stats.get(day, 0)
            percentage = (count / total_wins * 100) if total_wins > 0 else 0.0
            lines.append(f"{day_names[day]} ({day_labels[day]}): {percentage:.1f}% ({count}次)")
        
        return "\n".join(lines)
    
    # 準備寫入 Excel 的資料（水平排列）
    rows = []
    
    for idx, strategy in enumerate(best_strategies[:5], 1):
        ball_a_name = ball_names[strategy['ball_a_index']]
        ball_b_name = ball_names[strategy['ball_b_index']]
        missed_weeks_str = format_missed_weeks(strategy.get('missed_weeks', []))
        day_stats_str = format_day_stats(strategy.get('day_stats', {}), strategy.get('wins', 0))
        
        # 號碼組格式：第X顆球+偏移量 第Y顆球+偏移量
        number_group = f"{ball_a_name}+{strategy['offset_a']} {ball_b_name}+{strategy['offset_b']}"
        # 勝率格式：XX%
        win_rate_str = f"{strategy['win_rate']:.0f}%"
        
        rows.append({
            '組別': f'第{idx}組',
            '號碼組': number_group,
            '勝率': win_rate_str,
            '槓龜週': missed_weeks_str,
            '每日中獎統計': day_stats_str
        })
        
        # 控制台輸出
        print(f"{'🏆' if idx == 1 else '🥈' if idx == 2 else '🥉' if idx == 3 else '🏅'} 第{idx}組最佳策略: {number_group}, 勝率={strategy['win_rate']:.1f}% (中獎: {strategy['wins']}/{strategy['total']})")
        print(f"   槓龜週: {missed_weeks_str}")
        print(f"   每日中獎統計:\n   {day_stats_str.replace(chr(10), chr(10) + '   ')}")
    
    # 如果策略不足5組，補齊空行
    while len(rows) < 5:
        idx = len(rows) + 1
        rows.append({
            '組別': f'第{idx}組',
            '號碼組': '無符合策略',
            '勝率': '無符合策略',
            '槓龜週': '無符合策略',
            '每日中獎統計': '無符合策略'
        })
    
    strategy_df = pd.DataFrame(rows)
    
    # 使用 openpyxl 來處理 Excel（保留原有分頁）
    try:
        book = load_workbook(file_path)
        
        # 如果 Monday_Strategy 分頁已存在，刪除它
        if 'Monday_Strategy' in book.sheetnames:
            print("🔄 刪除舊的 Monday_Strategy 分頁...")
            del book['Monday_Strategy']
        
        # 創建新的分頁
        from openpyxl.utils.dataframe import dataframe_to_rows
        ws = book.create_sheet('Monday_Strategy')
        
        # 寫入標題（加粗）
        from openpyxl.styles import Font, Alignment
        headers = ['組別', '號碼組', '勝率', '槓龜週', '每日中獎統計']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
        
        # 寫入資料
        for r_idx, row in enumerate(dataframe_to_rows(strategy_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # 設定自動換行（特別是「每日中獎統計」和「槓龜週」欄位）
                if c_idx >= 4:  # 槓龜週和每日中獎統計欄位
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 12  # 組別
        ws.column_dimensions['B'].width = 30  # 號碼組
        ws.column_dimensions['C'].width = 10  # 勝率
        ws.column_dimensions['D'].width = 30  # 槓龜週
        ws.column_dimensions['E'].width = 50  # 每日中獎統計
        
        # 保存
        book.save(file_path)
        print(f"✅ 成功將策略分析結果寫入 {file_path} 的 Monday_Strategy 分頁")
        return True
        
    except Exception as e:
        print(f"⚠️ 使用 openpyxl 寫入失敗: {e}")
        print("   嘗試使用 pandas 方法...")
        # 如果 openpyxl 方法失敗，嘗試使用 pandas
        try:
            # 先讀取原有資料
            existing_sheets = {}
            try:
                excel_file = pd.ExcelFile(file_path, engine='openpyxl')
                for sheet_name in excel_file.sheet_names:
                    if sheet_name != 'Monday_Strategy':
                        existing_sheets[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            except:
                pass
            
            # 寫入所有分頁（包括新的 Monday_Strategy）
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                # 寫入原有分頁
                for sheet_name, sheet_df in existing_sheets.items():
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                # 寫入新的策略分頁
                strategy_df.to_excel(writer, sheet_name='Monday_Strategy', index=False)
            
            print(f"✅ 成功將策略分析結果寫入 {file_path} 的 Monday_Strategy 分頁（使用 pandas）")
            return True
        except Exception as e2:
            print(f"❌ 使用 pandas 寫入也失敗: {e2}")
            import traceback
            traceback.print_exc()
            return False

def main():
    parser = argparse.ArgumentParser(description='週一冠軍策略分析')
    parser.add_argument('--type', choices=['539', 'fantasy5'], required=True, help='彩球類型')
    parser.add_argument('--file', required=True, help='Excel 檔案路徑')
    
    args = parser.parse_args()
    
    lottery_type = args.type
    file_path = args.file
    
    success = add_strategy_sheet(file_path, lottery_type)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()

